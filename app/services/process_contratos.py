import pandas as pd
import numpy as np
import io
import os
import re
import logging
import unicodedata
from typing import List, Optional
from pathlib import Path
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.utils import get_column_letter

# Configurar logging para debug
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)


# ========================================
# NOMES PADRONIZADOS DAS ABAS
# ========================================
def get_sheet_names(bank_type: str) -> dict:
    """
    Retorna os nomes padronizados das abas baseado no banco.
    """
    if bank_type == 'bemge':
        return {
            '3026-11': 'Bemge 3026-11-Habil.Não Homol.',
            '3026-12-AUD': 'Bemge 3026-12-Homol.Auditados',
            '3026-12-NAUD': 'Bemge 3026-12-Homol.Não Auditados',
            '3026-12-TODOS': 'Bemge 3026-12-Homol.Todos',
            '3026-12-ULTIMOS_TODOS': 'Bemge 3026-12-Últ2M.Todos',
            '3026-12-ULTIMOS_AUD': 'Bemge 3026-12-Últ2M.Auditados',
            '3026-12-ULTIMOS_NAUD': 'Bemge 3026-12-Últ2M.Não Auditados',
            '3026-15': 'Bemge 3026-15-Homol.Neg.Cob'
        }
    else:  # minas_caixa
        return {
            '3026-11': 'Minas Caixa 3026-11-Habil.Não Homol',
            '3026-12-AUD': 'Minas Caixa 3026-12-Homol. Auditado',
            '3026-12-NAUD': 'Minas Caixa 3026-12-Homol.Não Auditado',
            '3026-12-TODOS': 'Minas Caixa 3026-12-Homol.Todos',
            '3026-12-ULTIMOS_TODOS': 'Minas Caixa 3026-12-Últ2M.Todos',
            '3026-12-ULTIMOS_AUD': 'Minas Caixa 3026-12-Últ2M.Auditado',
            '3026-12-ULTIMOS_NAUD': 'Minas Caixa 3026-12-Últ2M.Não Audit.',
            '3026-15': 'Minas Caixa 3026-15-Homol.Neg.Cob'
        }


def detect_file_type(filename: str) -> str:
    """
    Detecta o tipo de arquivo baseado no nome.
    Retorna: '3026-11', '3026-12' ou '3026-15'
    """
    filename_upper = filename.upper()
    if '3026-11' in filename_upper or '302611' in filename_upper:
        return '3026-11'
    elif '3026-12' in filename_upper or '302612' in filename_upper:
        return '3026-12'
    elif '3026-15' in filename_upper or '302615' in filename_upper:
        return '3026-15'
    else:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo de arquivo não reconhecido: {filename}. Esperado: 3026-11, 3026-12 ou 3026-15"
        )


def get_bank_name(bank_type: str) -> str:
    """Converte bank_type para nome do banco"""
    bank_names = {
        'bemge': 'BEMGE',
        'minas_caixa': 'MINAS CAIXA'
    }
    return bank_names.get(bank_type, bank_type.upper())


def _normalize_bank_type_key(bank_type: str) -> str:
    if not bank_type:
        return ''
    bt = bank_type.strip().lower().replace(' ', '_').replace('-', '_')
    if bt in ('minascaixa', 'minas', 'mc', 'caixa_minas', 'caixa'):
        return 'minas_caixa'
    if bt in ('bem_ge',):
        return 'bemge'
    return bt


def _cell_id_string(x) -> str:
    """Converte célula (contrato / ID numérico do Excel) em string sem perder dígitos por float."""
    if pd.isna(x):
        return ''
    if isinstance(x, str):
        s = x.strip()
        if len(s) > 2 and s.endswith('.0') and s[:-2].lstrip('-').isdigit():
            return s[:-2]
        return s
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, float):
        if not np.isfinite(x):
            return ''
        xr = round(x)
        if abs(x - xr) < 1e-6:
            if abs(xr) >= 1e15:
                return f'{x:.0f}'
            return str(int(xr))
        s = str(x).strip().rstrip('0').rstrip('.')
        return s
    return str(x).strip()


def format_contrato_column(df: pd.DataFrame) -> pd.DataFrame:
    """
    Formata a coluna CONTRATO como texto, preservando zeros à esquerda.
    """
    if 'CONTRATO' in df.columns:
        df['CONTRATO'] = df['CONTRATO'].apply(_cell_id_string)
    return df


def format_column_d_as_text(df: pd.DataFrame) -> pd.DataFrame:
    """
    Formata a coluna D (índice 3) como texto.
    """
    if len(df.columns) > 3:
        coluna_d = df.columns[3]
        df[coluna_d] = df[coluna_d].apply(_cell_id_string)
    return df


def _stats_total_unicos(stats: dict, subset_key: str) -> int:
    """Lê total_unicos com fallbacks (evita KeyError auditados/nauditados em integrações antigas)."""
    blk = stats.get(subset_key)
    if blk is None and subset_key == 'aud':
        blk = stats.get('auditados')
    if blk is None and subset_key == 'naud':
        blk = stats.get('nauditados') or stats.get('Nauditados')
    if isinstance(blk, dict):
        return int(blk.get('total_unicos', 0))
    return 0


def _normalize_auditado_token(v) -> str:
    if pd.isna(v):
        return ''
    t = str(v).strip().upper()
    t = unicodedata.normalize('NFKD', t)
    t = ''.join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r'\s+', '', t)
    return t


def resolve_manifestacao_column(
    df: pd.DataFrame,
    bank_type: Optional[str] = None,
    date_column: Optional[str] = None
) -> Optional[str]:
    """
    Descobre coluna de data para filtro de período (manifestação ou equivalente).
    Varre índices típicos e, em planilhas largas, faixa 20–54 para achar a coluna com mais datas válidas.
    """
    if df.empty:
        return None
    if date_column and date_column in df.columns:
        return date_column

    possible_cols = [
        'DT.MANIFESTACAO', 'DT.MANIFESTAÇÃO', 'DATA MANIFESTACAO', 'DATA MANIFESTAÇÃO',
    ]
    for col in possible_cols:
        if col in df.columns:
            return col

    for col in df.columns:
        if 'MANIFEST' in str(col).upper():
            return col

    candidate_indices = [32, 31, 33, 30, 34, 29, 28, 35, 26, 27, 36]
    if bank_type == 'minas_caixa':
        candidate_indices = [32, 31, 33, 30, 34, 29, 28, 24] + [19, 23, 25]
    wide_scan = list(range(20, min(55, len(df.columns))))
    candidate_indices = list(dict.fromkeys(candidate_indices + wide_scan))

    best_col, best_score = None, 0.0
    for idx in candidate_indices:
        if len(df.columns) <= idx:
            continue
        col = df.columns[idx]
        s = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
        score = float(s.notna().mean())
        if score > best_score:
            best_col, best_score = col, score

    min_ratio = 0.035
    if best_col is not None and best_score >= min_ratio:
        logger.info(f"Coluna de período inferida ({best_score:.0%} válida): {best_col!r}")
        return best_col

    if bank_type != 'minas_caixa' and len(df.columns) > 32:
        logger.debug(f"Fallback período coluna índice 32: {df.columns[32]!r}")
        return df.columns[32]

    logger.warning(f"Não foi possível inferir coluna de manifestação/período (bank_type={bank_type})")
    return None


def format_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Formata colunas de data, removendo a hora.
    Colunas afetadas: DT.ASS., DT.EVENTO, DT.HAB., DT.PROC.HAB. e variações
    """
    # Lista de possíveis nomes de colunas de data
    colunas_data = [
        'DT.ASS.', 'DT.EVENTO', 'DT.HAB.', 'DT.PROC.HAB.',
        'DT.ASS', 'DT.EVENTO', 'DT.HAB', 'DT.PROC.HAB',
        'DATA ASS.', 'DATA EVENTO', 'DATA HAB.', 'DATA PROC.HAB.',
        'DT.BASE', 'DT.TERM.ANALISE', 'DT.MANIFESTACAO', 'DT.POS.NOVACAO',
        'DT.ULT.AUDITORIA', 'DT.ULT.NEGOCIACAO', 'DATA STATUS'
    ]
    
    for col in df.columns:
        # Verificar se a coluna está na lista ou começa com DT. ou DATA
        if col in colunas_data or col.upper().startswith('DT.') or col.upper().startswith('DATA'):
            try:
                # Converter para datetime e extrair apenas a data
                df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True).dt.date
                logger.debug(f"Coluna de data formatada: {col}")
            except Exception as e:
                logger.warning(f"Erro ao formatar coluna de data {col}: {e}")
    
    return df


def format_object_columns_that_look_like_dates(df: pd.DataFrame, max_scan_cols: int = 80) -> pd.DataFrame:
    """
    Converte colunas object que na maior parte são datas (evita horas visíveis em abas Minas/Bemge).
    """
    if df.empty:
        return df
    df = df.copy()
    n = min(len(df.columns), max_scan_cols)
    for i, col in enumerate(df.columns[:n]):
        if col in ('BANCO', 'TIPO_ARQUIVO', 'AUDITADO_TIPO'):
            continue
        ser = df[col]
        if ser.dtype != object:
            continue
        sample = ser.dropna().head(400)
        if len(sample) < 5:
            continue
        parsed = pd.to_datetime(sample, errors='coerce', dayfirst=True)
        ratio = float(parsed.notna().mean())
        if ratio < 0.45:
            continue
        try:
            df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True).dt.date
            logger.debug(f"Coluna object tratada como data: {col!r} (~{ratio:.0%} amostra válida)")
        except Exception:
            continue
    return df


def format_date_columns_by_index(df: pd.DataFrame, indices: list) -> pd.DataFrame:
    """
    Formata colunas de data específicas por índice (0-based), removendo a hora.
    Usado para MINAS CAIXA 3026-11 colunas T, X, Z (índices 19, 23, 25).
    """
    for idx in indices:
        if len(df.columns) > idx:
            col = df.columns[idx]
            try:
                df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True).dt.date
                logger.debug(f"Coluna de data (índice {idx}) formatada: {col}")
            except Exception as e:
                logger.warning(f"Erro ao formatar coluna de data índice {idx} ({col}): {e}")
    
    return df


def remove_general_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove colunas AF, AG, AH (INDVAF3TR7, INDVAF4TR7, DT.ULT.HOMOLOGACAO).
    """
    colunas_remover_geral = ['INDVAF3TR7', 'INDVAF4TR7', 'DT.ULT.HOMOLOGACAO']
    for col in colunas_remover_geral:
        if col in df.columns:
            df = df.drop(columns=[col])
    return df


# ========================================
# FILTRO HABITACIONAL (coluna Y — BEMGE só Y; Minas Y)
# ========================================
def apply_habitacional_filter(
    df: pd.DataFrame,
    bank_type: str,
    reference_date: Optional[str] = None,
    months_back: int = 2
) -> pd.DataFrame:
    """
    Filtro habitacional por data na coluna Y (índice 24), janela [corte, referência].
    BEMGE: somente coluna Y (definição operacional dos testes).
    MINAS CAIXA: coluna Y (índice 24).
    """
    if df.empty or not reference_date:
        logger.debug("Filtro habitacional: DataFrame vazio ou sem data de referência")
        return df
    
    initial_count = len(df)
    logger.info(f"🔍 APLICANDO FILTRO HABITACIONAL ({bank_type.upper()})")
    logger.info(f"   Registros iniciais: {initial_count}")
    logger.info(f"   Data referência: {reference_date}")
    logger.info(f"   Meses atrás: {months_back}")
    
    try:
        # Calcular intervalo de datas
        data_ref = datetime.strptime(reference_date, "%Y-%m-%d").date()
        data_corte = data_ref - relativedelta(months=months_back)
        
        logger.info(f"   Intervalo: {data_corte} até {data_ref}")
        
        columns_to_check = []
        if bank_type == 'bemge':
            logger.info("   BEMGE: filtro habitacional apenas na coluna Y (índice 24)")
            if len(df.columns) > 24:
                columns_to_check.append((24, 'Y'))
        elif bank_type == 'minas_caixa':
            logger.info("   MINAS CAIXA: coluna Y (índice 24)")
            if len(df.columns) > 24:
                columns_to_check.append((24, 'Y'))
        
        if not columns_to_check:
            logger.warning(f"   ⚠️  Colunas habitacionais não encontradas!")
            logger.warning(f"   Total de colunas no DataFrame: {len(df.columns)}")
            return df
        
        # Criar máscara combinada (OR lógico entre colunas)
        combined_mask = pd.Series(False, index=df.index)
        found_valid_data = False
        
        for col_idx, col_name in columns_to_check:
            col = df.columns[col_idx]
            logger.info(f"   Processando coluna {col_name} (índice {col_idx}): '{col}'")
            
            # Tentar converter para datetime
            parsed_dates = pd.to_datetime(df[col], errors='coerce', dayfirst=True)
            valid_dates = parsed_dates.notna().sum()
            
            logger.info(f"      Datas válidas: {valid_dates}/{initial_count}")
            
            if valid_dates == 0:
                logger.warning(f"      ⚠️  Nenhuma data válida nesta coluna")
                continue
            
            found_valid_data = True
            
            # Mostrar range de datas
            min_date = parsed_dates.min()
            max_date = parsed_dates.max()
            logger.info(f"      Range de datas: {min_date.date()} até {max_date.date()}")
            
            # Criar máscara para esta coluna
            mask = (
                parsed_dates.notna()
                & (parsed_dates >= pd.Timestamp(data_corte))
                & (parsed_dates <= pd.Timestamp(data_ref))
            )
            
            # OR lógico com máscara combinada
            combined_mask |= mask
            
            logger.info(f"      ✅ {mask.sum()} registros no intervalo nesta coluna")
        
        if not found_valid_data:
            logger.warning(f"   ⚠️  Nenhuma coluna com datas válidas encontrada")
            return df
        
        # Aplicar máscara combinada
        df_filtrado = df[combined_mask].copy()
        result_count = len(df_filtrado)
        
        logger.info(f"   ✅ FILTRO HABITACIONAL APLICADO")
        logger.info(f"   Registros que passaram: {result_count}/{initial_count}")
        logger.info(f"   Registros removidos: {initial_count - result_count}")
        
        if result_count == 0:
            logger.warning(f"   ⚠️  WARNING: Nenhum registro passou no filtro habitacional!")
            logger.warning(f"   Verifique se a data de referência está correta")
        
        return df_filtrado
        
    except Exception as e:
        logger.error(f"   ❌ ERRO no filtro habitacional: {e}")
        logger.error(f"   Retornando dados originais")
        return df


def filter_by_period(
    df: pd.DataFrame, 
    reference_date: str = None, 
    months_back: int = 2,
    date_column: str = None,
    bank_type: Optional[str] = None
) -> pd.DataFrame:
    """
    Filtra contratos por período (manifestação ou coluna inferida).
    Janela: [data_corte, data_referência] (inclusive), com data_corte = ref - months_back.
    """
    if df.empty:
        return df
    
    initial_count = len(df)
    logger.info(f"🔍 APLICANDO FILTRO DE PERÍODO")
    logger.info(f"   Registros iniciais: {initial_count}")
    logger.info(f"   Data referência: {reference_date or 'Data atual'}")
    logger.info(f"   Meses atrás: {months_back}")
    
    col_manifestacao = resolve_manifestacao_column(df, bank_type=bank_type, date_column=date_column)
    
    if col_manifestacao is None:
        logger.warning("   ⚠️  WARNING: Coluna de data de manifestação não encontrada")
        logger.warning(f"   Colunas disponíveis: {list(df.columns[:15])}...")
        logger.warning("   ❌ Filtro de período NÃO aplicado")
        return df
    
    logger.info(f"   ✅ Coluna de período encontrada: '{col_manifestacao}'")
    
    try:
        # Determinar data de referência
        if reference_date:
            try:
                data_ref = datetime.strptime(reference_date, "%Y-%m-%d").date()
            except ValueError:
                logger.error(f"Formato de data inválido: {reference_date}. Esperado: YYYY-MM-DD")
                data_ref = datetime.now().date()
        else:
            data_ref = datetime.now().date()
        
        # Calcular data de corte
        data_corte = data_ref - relativedelta(months=months_back)
        
        logger.info(f"   Intervalo: {data_corte} até {data_ref}")
        
        df_temp = df.copy()
        parsed = pd.to_datetime(df_temp[col_manifestacao], errors='coerce', dayfirst=True)
        parsed_day = parsed.dt.floor('D')
        
        valid_dates = int(parsed.notna().sum())
        logger.info(f"   Datas válidas: {valid_dates}/{initial_count}")
        
        if valid_dates == 0:
            logger.warning("   ⚠️  WARNING: Nenhuma data válida encontrada")
            logger.warning("   ❌ Filtro de período NÃO aplicado")
            return df
        
        try:
            min_date = parsed.min()
            max_date = parsed.max()
            if pd.notna(min_date) and pd.notna(max_date):
                logger.info(
                    f"   Range de datas nos dados: {pd.Timestamp(min_date).date()} até {pd.Timestamp(max_date).date()}"
                )
        except Exception as ex:
            logger.warning(f"   Não foi possível exibir range de datas: {ex}")
        
        mask = (parsed_day >= pd.Timestamp(data_corte)) & (parsed_day <= pd.Timestamp(data_ref))
        df_filtrado = df[mask].copy()
        
        result_count = len(df_filtrado)
        logger.info(f"   ✅ FILTRO DE PERÍODO APLICADO")
        logger.info(f"   Registros que passaram: {result_count}/{initial_count}")
        logger.info(f"   Registros removidos: {initial_count - result_count}")
        
        if result_count == 0:
            logger.warning("   ⚠️  WARNING: Nenhum registro no intervalo especificado!")
            logger.warning("   Verifique se a data de referência está correta")
        
        return df_filtrado
        
    except Exception as e:
        logger.error(f"   ❌ ERRO ao filtrar por período: {e}")
        logger.error("   Retornando dados originais")
        return df


def filter_last_2_months(df: pd.DataFrame, date_column: str = None) -> pd.DataFrame:
    """
    Filtra contratos dos últimos 2 meses baseado na coluna de data de manifestação.
    Função de compatibilidade - usa filter_by_period internamente.
    """
    return filter_by_period(df, reference_date=None, months_back=2, date_column=date_column)


def process_3026_11(df: pd.DataFrame, bank_name: str) -> tuple:
    """
    Processa arquivos 3026-11.
    - Formata coluna D como texto
    - Remove duplicados na coluna CONTRATO
    - Para MINAS CAIXA: formata colunas T, X, Z (índices 19, 23, 25) removendo horas
    Retorna: (df_processado, total_linhas, total_unicos, total_duplicados)
    """
    logger.debug(f"Processando 3026-11 para {bank_name} - Linhas: {len(df)}, Colunas: {len(df.columns)}")
    logger.debug(f"Colunas do 3026-11: {list(df.columns)[:10]}...")
    
    # Verificar se DataFrame está vazio
    if df.empty:
        logger.warning(f"DataFrame 3026-11 está vazio para {bank_name}")
        return pd.DataFrame(), 0, 0, 0
    
    # Fazer cópia para evitar problemas de referência
    df = df.copy()
    
    # MINAS CAIXA: datas em T, X, Z e colunas comuns de manifestação / período
    if 'MINAS' in bank_name.upper():
        logger.debug("Formatação de datas (3026-11 Minas: T,X,Z e índices de período)")
        df = format_date_columns_by_index(df, [19, 23, 25, 24, 31, 32, 33, 30])
    
    # Formatar coluna D como texto
    df = format_column_d_as_text(df)
    
    # Verificar se tem coluna CONTRATO (tentar variações)
    coluna_contrato = None
    for nome in ['CONTRATO', 'CONTRATO_NUM', 'NUM_CONTRATO', 'NR_CONTRATO']:
        if nome in df.columns:
            coluna_contrato = nome
            break
    
    if coluna_contrato is None:
        logger.error(f"Coluna CONTRATO não encontrada. Colunas disponíveis: {list(df.columns)}")
        raise HTTPException(
            status_code=400,
            detail="Coluna 'CONTRATO' não encontrada no arquivo 3026-11"
        )
    
    # Renomear para CONTRATO se necessário
    if coluna_contrato != 'CONTRATO':
        df = df.rename(columns={coluna_contrato: 'CONTRATO'})
        logger.debug(f"Coluna {coluna_contrato} renomeada para CONTRATO")

    df = format_contrato_column(df)
    
    total_linhas = len(df)
    nunique_c = int(df['CONTRATO'].nunique())

    # Minas: não remover linhas com mesmo contrato (evita “sumir” contrato por arredondamento Excel)
    if 'MINAS' in bank_name.upper():
        df_processado = df.copy()
        total_unicos = nunique_c
        total_duplicados = total_linhas - nunique_c
        logger.debug(f"3026-11 Minas: sem drop_duplicates por CONTRATO; linhas={total_linhas}, nunique={nunique_c}")
    else:
        df_processado = df.drop_duplicates(subset=['CONTRATO'], keep='first').copy()
        total_unicos = len(df_processado)
        total_duplicados = total_linhas - total_unicos
    
    logger.debug(f"3026-11 processado: total={total_linhas}, únicos={total_unicos}, duplicados={total_duplicados}")
    
    return df_processado, total_linhas, total_unicos, total_duplicados


def process_3026_15(df: pd.DataFrame, bank_name: str) -> tuple:
    """
    Processa arquivos 3026-15.
    - Formata coluna D como texto
    - Remove duplicados na coluna CONTRATO
    Retorna: (df_processado, total_linhas, total_unicos, total_duplicados)
    """
    logger.debug(f"Processando 3026-15 para {bank_name} - Linhas: {len(df)}, Colunas: {len(df.columns)}")
    logger.debug(f"Colunas do 3026-15: {list(df.columns)[:10]}...")
    
    # Verificar se DataFrame está vazio
    if df.empty:
        logger.warning(f"DataFrame 3026-15 está vazio para {bank_name}")
        return pd.DataFrame(), 0, 0, 0
    
    # Fazer cópia para evitar problemas de referência
    df = df.copy()
    
    # Formatar coluna D como texto
    df = format_column_d_as_text(df)
    
    # Verificar se tem coluna CONTRATO (tentar variações)
    coluna_contrato = None
    for nome in ['CONTRATO', 'CONTRATO_NUM', 'NUM_CONTRATO', 'NR_CONTRATO']:
        if nome in df.columns:
            coluna_contrato = nome
            break
    
    if coluna_contrato is None:
        logger.error(f"Coluna CONTRATO não encontrada. Colunas disponíveis: {list(df.columns)}")
        raise HTTPException(
            status_code=400,
            detail="Coluna 'CONTRATO' não encontrada no arquivo 3026-15"
        )
    
    # Renomear para CONTRATO se necessário
    if coluna_contrato != 'CONTRATO':
        df = df.rename(columns={coluna_contrato: 'CONTRATO'})
        logger.debug(f"Coluna {coluna_contrato} renomeada para CONTRATO")

    df = format_contrato_column(df)
    
    total_linhas = len(df)
    nunique_c = int(df['CONTRATO'].nunique())

    if 'MINAS' in bank_name.upper():
        df_processado = df.copy()
        total_unicos = nunique_c
        total_duplicados = total_linhas - nunique_c
        logger.debug(f"3026-15 Minas: sem drop_duplicates por CONTRATO")
    else:
        df_processado = df.drop_duplicates(subset=['CONTRATO'], keep='first').copy()
        total_unicos = len(df_processado)
        total_duplicados = total_linhas - total_unicos
    
    logger.debug(f"3026-15 processado: total={total_linhas}, únicos={total_unicos}, duplicados={total_duplicados}")
    
    return df_processado, total_linhas, total_unicos, total_duplicados


def process_3026_12(df: pd.DataFrame, bank_name: str) -> dict:
    """
    Processa arquivo 3026-12.
    - Coluna B: Manter apenas linhas onde = 52101
    - Coluna D: Formatar como texto
    - Colunas AA e AB: Deixar como número único (sem decimais)
    - Remove colunas BT e BU
    - Separa em AUD e NAUD, aplica filtros
    Retorna: {
        'aud': (df_aud, total_aud, unicos_aud, duplicados_aud),
        'naud': (df_naud, total_naud, unicos_naud, duplicados_naud)
    }
    """
    logger.debug(f"Processando 3026-12 para {bank_name} - Colunas: {len(df.columns)}, Linhas: {len(df)}")
    
    # Verificar se DataFrame está vazio
    if df.empty:
        logger.warning(f"DataFrame 3026-12 está vazio para {bank_name}")
        return {
            'aud': (pd.DataFrame(), 0, 0, 0),
            'naud': (pd.DataFrame(), 0, 0, 0),
            'todos_full': pd.DataFrame(),
        }
    
    # Fazer cópia para evitar problemas de referência
    df = df.copy()
    
    logger.debug(f"Colunas do 3026-12: {list(df.columns)[:15]}...")
    
    # 1. Coluna B - Manter apenas linhas onde = 52101
    if len(df.columns) > 1:
        coluna_b = df.columns[1]
        logger.debug(f"Coluna B (índice 1): {coluna_b}")
        # Converter para string, removendo .0 de números float
        df[coluna_b] = df[coluna_b].apply(_cell_id_string)
        df[coluna_b] = df[coluna_b].str.strip()
        linhas_antes = len(df)
        
        # Verificar se existe o valor 52101
        valores_unicos = df[coluna_b].unique()[:10]
        logger.debug(f"Valores únicos na coluna B (amostra): {valores_unicos}")
        
        # Filtrar por 52101
        df_filtrado = df[df[coluna_b] == '52101'].copy()
        linhas_depois = len(df_filtrado)
        logger.debug(f"Filtro coluna B=52101: {linhas_antes} -> {linhas_depois} linhas")
        
        # Se não encontrou nenhuma linha, continuar sem o filtro
        if linhas_depois == 0:
            logger.warning(f"Nenhuma linha com valor 52101 encontrada. Continuando sem filtro da coluna B.")
        else:
            df = df_filtrado
    
    # Se não sobrou nenhuma linha
    if len(df) == 0:
        logger.warning("DataFrame vazio após processamento inicial")
        return {
            'aud': (pd.DataFrame(), 0, 0, 0),
            'naud': (pd.DataFrame(), 0, 0, 0),
            'todos_full': pd.DataFrame()
        }
    
    # 2. Formatar coluna D como texto
    df = format_column_d_as_text(df)
    
    # 3. Colunas AA e AB - deixar como número único (sem decimais)
    for idx in [26, 27]:  # AA=26, AB=27 (0-based)
        if len(df.columns) > idx:
            col = df.columns[idx]
            logger.debug(f"Formatando coluna índice {idx} ({col}) como inteiro")
            df[col] = df[col].apply(
                lambda x: int(x) if pd.notna(x) and isinstance(x, (int, float)) else x
            )
    
    # 4. Remover colunas BT e BU (índices 71 e 72, 0-based)
    if len(df.columns) > 72:
        colunas_bt_bu = df.columns[71:73].tolist()
        logger.debug(f"Removendo colunas BT e BU: {colunas_bt_bu}")
        df = df.drop(columns=colunas_bt_bu, errors='ignore')
    
    # Verificar coluna AUDITADO (tentar diferentes nomes)
    coluna_auditado = None
    for nome in ['AUDITADO', 'AUD', 'AUDIT']:
        if nome in df.columns:
            coluna_auditado = nome
            break
    
    if coluna_auditado is None:
        logger.error(f"Coluna AUDITADO não encontrada. Colunas disponíveis: {list(df.columns)}")
        raise HTTPException(
            status_code=400,
            detail=f"Coluna 'AUDITADO' não encontrada no arquivo 3026-12"
        )
    
    # Verificar coluna CONTRATO (tentar variações)
    coluna_contrato = None
    for nome in ['CONTRATO', 'CONTRATO_NUM', 'NUM_CONTRATO', 'NR_CONTRATO']:
        if nome in df.columns:
            coluna_contrato = nome
            break
    
    if coluna_contrato is None:
        logger.error(f"Coluna CONTRATO não encontrada. Colunas disponíveis: {list(df.columns)}")
        raise HTTPException(
            status_code=400,
            detail=f"Coluna 'CONTRATO' não encontrada no arquivo 3026-12"
        )
    
    # Renomear para CONTRATO se necessário
    if coluna_contrato != 'CONTRATO':
        df = df.rename(columns={coluna_contrato: 'CONTRATO'})
        logger.debug(f"Coluna {coluna_contrato} renomeada para CONTRATO")
    
    # Verificar colunas de filtro (nomes reais das colunas)
    filter_cols = ['DEST.PAGAM', 'DEST.COMPLEM']
    has_filter_cols = any(col in df.columns for col in filter_cols)
    
    # Tokens já no formato de _normalize_auditado_token (sem espaços / sem acentos)
    aud_tokens = {
        'AUDI', 'AUD', 'AUDITADO', 'AUDITADOS', 'AUDIT.', 'SIM', 'S',
        '1', '1.0', 'TRUE', 'VERDADEIRO',
    }
    naud_tokens = {
        'NAUD', 'NAUD.', 'NAOAUDITADO', 'NAUDITADO', 'NAOAUDITADOS', 'NAUDITADOS',
        'NAO', '2', '2.0', 'FALSE', 'FALSO', 'NAOAUD.',
    }
    keys = df[coluna_auditado].map(_normalize_auditado_token)
    extra_map = {_normalize_auditado_token(v) for v in df[coluna_auditado].unique() if pd.notna(v)}
    logger.debug(f"AUDITADO chaves normalizadas (amostra): {list(extra_map)[:25]}")
    
    df_aud = df[keys.isin(aud_tokens)].copy()
    df_naud = df[keys.isin(naud_tokens)].copy()
    
    logger.debug(f"Após separação: AUD={len(df_aud)}, NAUD={len(df_naud)}")
    
    classified = keys.isin(aud_tokens) | keys.isin(naud_tokens)
    n_unc = int((~classified).sum())
    if n_unc:
        amostra = df.loc[~classified, coluna_auditado].dropna().unique()[:8]
        logger.warning(f"3026-12: {n_unc} linhas com AUDITADO não classificado; valores: {amostra}")
    
    df_todos_full = df.copy()
    df_todos_full['BANCO'] = bank_name
    df_todos_full['TIPO_ARQUIVO'] = '3026-12'
    df_todos_full['AUDITADO_TIPO'] = keys.map(
        lambda k: 'AUD' if k in aud_tokens else ('NAUD' if k in naud_tokens else 'INDEF')
    )
    df_todos_full['DUPLICADO'] = df_todos_full['CONTRATO'].duplicated(keep=False)
    
    # Aplicar filtros se as colunas existirem
    valores_filtro = ['0X0', '1X4', '6X4', '8X4', '0x0', '1x4', '6x4', '8x4']
    
    if has_filter_cols:
        for col in filter_cols:
            if col in df_aud.columns and len(df_aud) > 0:
                df_aud[col] = df_aud[col].astype(str).str.upper().str.strip()
                mask = ~df_aud[col].isin(valores_filtro)
                df_aud = df_aud[mask].copy()
            
            if col in df_naud.columns and len(df_naud) > 0:
                df_naud[col] = df_naud[col].astype(str).str.upper().str.strip()
                mask = ~df_naud[col].isin(valores_filtro)
                df_naud = df_naud[mask].copy()
    
    # Processar AUD
    total_aud = len(df_aud)
    if total_aud > 0:
        df_aud_unicos = df_aud.drop_duplicates(subset=['CONTRATO'], keep='first')
        unicos_aud = len(df_aud_unicos)
    else:
        df_aud_unicos = pd.DataFrame()
        unicos_aud = 0
    duplicados_aud = total_aud - unicos_aud
    
    # Processar NAUD
    total_naud = len(df_naud)
    if total_naud > 0:
        df_naud_unicos = df_naud.drop_duplicates(subset=['CONTRATO'], keep='first')
        unicos_naud = len(df_naud_unicos)
    else:
        df_naud_unicos = pd.DataFrame()
        unicos_naud = 0
    duplicados_naud = total_naud - unicos_naud
    
    logger.debug(f"AUD: total={total_aud}, únicos={unicos_aud}, duplicados={duplicados_aud}")
    logger.debug(f"NAUD: total={total_naud}, únicos={unicos_naud}, duplicados={duplicados_naud}")
    
    return {
        'aud': (df_aud_unicos, total_aud, unicos_aud, duplicados_aud),
        'naud': (df_naud_unicos, total_naud, unicos_naud, duplicados_naud),
        'todos_full': df_todos_full
    }


def filtrar_planilha_contratos(
    df: pd.DataFrame,
    aplicar_periodo: bool = False,
    reference_date: Optional[str] = None,
    months_back: int = 2,
    aplicar_habitacional: bool = False,
    aplicar_3026_15: bool = False,
    date_column: str = None,
    bank_type: str = None
) -> pd.DataFrame:
    """
    ✅ CORRIGIDO: Aplica filtros condicionais em um DataFrame de contratos.
    Agora inclui filtro habitacional (colunas W e Y).
    """
    if df.empty:
        return df

    df_filtrado = df.copy()

    # Filtro de período
    if aplicar_periodo:
        df_filtrado = filter_by_period(
            df_filtrado, reference_date, months_back, date_column, bank_type=bank_type
        )

    # ✅ NOVO: Filtro habitacional (colunas W e Y)
    if aplicar_habitacional and bank_type:
        df_filtrado = apply_habitacional_filter(
            df_filtrado,
            bank_type=bank_type,
            reference_date=reference_date,
            months_back=months_back
        )

    # DEST.PAGAM / DEST.COMPLEM: regra do 3026-12 — não aplicar em 3026-11 / 3026-15
    valores_filtro = {'0X0', '1X4', '6X4', '8X4'}
    filter_cols = ['DEST.PAGAM', 'DEST.COMPLEM']
    if 'TIPO_ARQUIVO' in df_filtrado.columns:
        mask_somente_12 = df_filtrado['TIPO_ARQUIVO'].eq('3026-12')
    else:
        mask_somente_12 = pd.Series(True, index=df_filtrado.index)
    for col in filter_cols:
        if col not in df_filtrado.columns or df_filtrado.empty:
            continue
        norm = df_filtrado[col].astype(str).str.upper().str.strip()
        remove = norm.isin(valores_filtro) & mask_somente_12
        df_filtrado = df_filtrado[~remove].copy()

    if aplicar_3026_15 and 'TIPO_ARQUIVO' in df_filtrado.columns:
        df_filtrado = df_filtrado[df_filtrado['TIPO_ARQUIVO'] == '3026-15'].copy()

    return df_filtrado


def processar_3026_12_com_abas(
    df: pd.DataFrame,
    bank_name: str,
    bank_type: str,
    period_filter_active: bool,
    reference_date: Optional[str],
    months_back: int
) -> dict:
    """
    ✅ CORRIGIDO: Processa 3026-12 com tratamento robusto de erros
    """
    try:
        resumo = process_3026_12(df, bank_name)

        if 'aud' not in resumo or 'naud' not in resumo or 'todos_full' not in resumo:
            logger.error(f"❌ Resumo de 3026-12 inválido. Chaves: {resumo.keys()}")
            raise ValueError("Estrutura de resumo inválida para 3026-12")

        df_aud, total_aud, unicos_aud, duplicados_aud = resumo['aud']
        df_naud, total_naud, unicos_naud, duplicados_naud = resumo['naud']
        df_todos = resumo['todos_full']

        logger.info(f"📊 3026-12 separado: AUD={len(df_aud)}, NAUD={len(df_naud)}, TODOS={len(df_todos)}")

        def preparar_sub_df(sub_df: pd.DataFrame, tipo: str) -> pd.DataFrame:
            if sub_df.empty:
                logger.debug(f"DataFrame vazio para tipo {tipo}")
                return pd.DataFrame()

            df_copy = sub_df.copy()
            df_copy['BANCO'] = bank_name
            df_copy['TIPO_ARQUIVO'] = '3026-12'
            df_copy['AUDITADO_TIPO'] = 'AUD' if tipo == 'aud' else 'NAUD'
            df_copy['DUPLICADO'] = df_copy['CONTRATO'].duplicated(keep=False)
            logger.debug(f"Sub-dataframe preparado ({tipo}): {len(df_copy)} registros")
            return df_copy

        df_aud_processado = preparar_sub_df(df_aud, 'aud')
        df_naud_processado = preparar_sub_df(df_naud, 'naud')

        # Criar dicionário de períodos
        periodos = {}
        
        if period_filter_active:
            logger.info(f"🔍 Aplicando filtro de período para 3026-12...")
            
            # Filtrar auditados
            if not df_aud_processado.empty:
                logger.info(f"   Filtrando AUD: {len(df_aud_processado)} registros iniciais")
                periodos['auditados_ultimos_2_meses'] = filtrar_planilha_contratos(
                    df_aud_processado,
                    aplicar_periodo=True,
                    reference_date=reference_date,
                    months_back=months_back,
                    bank_type=bank_type
                )
                logger.info(f"   AUD filtrado: {len(periodos['auditados_ultimos_2_meses'])} registros")
            else:
                periodos['auditados_ultimos_2_meses'] = pd.DataFrame()
            
            # Filtrar não auditados
            if not df_naud_processado.empty:
                logger.info(f"   Filtrando NAUD: {len(df_naud_processado)} registros iniciais")
                periodos['naud_ultimos_2_meses'] = filtrar_planilha_contratos(
                    df_naud_processado,
                    aplicar_periodo=True,
                    reference_date=reference_date,
                    months_back=months_back,
                    bank_type=bank_type
                )
                logger.info(f"   NAUD filtrado: {len(periodos['naud_ultimos_2_meses'])} registros")
            else:
                periodos['naud_ultimos_2_meses'] = pd.DataFrame()
            
            # Filtrar todos
            if not df_todos.empty:
                logger.info(f"   Filtrando TODOS: {len(df_todos)} registros iniciais")
                periodos['todos_ultimos_2_meses'] = filtrar_planilha_contratos(
                    df_todos,
                    aplicar_periodo=True,
                    reference_date=reference_date,
                    months_back=months_back,
                    bank_type=bank_type
                )
                logger.info(f"   TODOS filtrado: {len(periodos['todos_ultimos_2_meses'])} registros")
            else:
                periodos['todos_ultimos_2_meses'] = pd.DataFrame()
        else:
            logger.info("ℹ️  Filtro de período desativado para 3026-12")
            periodos['auditados_ultimos_2_meses'] = pd.DataFrame()
            periodos['naud_ultimos_2_meses'] = pd.DataFrame()
            periodos['todos_ultimos_2_meses'] = pd.DataFrame()

        logger.info("✅ Abas 3026-12 construídas com sucesso")

        st_aud = {
            'total_linhas': total_aud,
            'total_unicos': unicos_aud,
            'total_duplicados': duplicados_aud
        }
        st_naud = {
            'total_linhas': total_naud,
            'total_unicos': unicos_naud,
            'total_duplicados': duplicados_naud
        }
        abas_out = {
            'todos': df_todos,
            'aud': df_aud_processado,
            'naud': df_naud_processado,
            'auditados_ultimos_2_meses': periodos['auditados_ultimos_2_meses'],
            'naud_ultimos_2_meses': periodos['naud_ultimos_2_meses'],
            'todos_ultimos_2_meses': periodos['todos_ultimos_2_meses'],
        }
        abas_out['auditados'] = df_aud_processado
        abas_out['nauditados'] = df_naud_processado
        return {
            'abas': abas_out,
            'stats': {
                'aud': st_aud,
                'naud': st_naud,
                'auditados': st_aud,
                'nauditados': st_naud,
                'Nauditados': st_naud,
            }
        }
    
    except Exception as e:
        logger.error(f"❌ ERRO em processar_3026_12_com_abas: {e}")
        import traceback
        logger.error(traceback.format_exc())
        raise


def gerar_resumo_geral(df_full: pd.DataFrame) -> pd.DataFrame:
    if df_full.empty:
        return pd.DataFrame()

    df = df_full.copy()
    if 'BANCO' in df.columns:
        df['BANCO'] = df['BANCO'].fillna('NÃO INFORMADO')
    else:
        df['BANCO'] = 'NÃO INFORMADO'
    if 'DUPLICADO' in df.columns:
        df['DUPLICADO'] = df['DUPLICADO'].fillna(False)
    else:
        df['DUPLICADO'] = False

    group_cols = ['BANCO', 'TIPO_ARQUIVO']
    if 'AUDITADO_TIPO' in df.columns:
        group_cols.append('AUDITADO_TIPO')

    summary = (
        df.groupby(group_cols, dropna=False)
        .agg(
            TOTAL_LINHAS=('CONTRATO', 'size'),
            CONTRATOS_UNICOS=('CONTRATO', lambda s: s.nunique()),
            CONTRATOS_DUPLICADOS=('DUPLICADO', lambda s: int(s.sum()) if not s.empty else 0)
        )
        .reset_index()
    )

    total_row = {
        'BANCO': 'TOTAL GERAL',
        'TIPO_ARQUIVO': '-',
        'TOTAL_LINHAS': summary['TOTAL_LINHAS'].sum(),
        'CONTRATOS_UNICOS': summary['CONTRATOS_UNICOS'].sum(),
        'CONTRATOS_DUPLICADOS': summary['CONTRATOS_DUPLICADOS'].sum()
    }
    if 'AUDITADO_TIPO' in summary.columns:
        total_row['AUDITADO_TIPO'] = '-'

    summary = pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)
    return summary


def gerar_contratos_repetidos(df_full: pd.DataFrame) -> pd.DataFrame:
    if df_full.empty:
        return pd.DataFrame()

    if 'DUPLICADO' in df_full.columns:
        df_repetidos = df_full[df_full['DUPLICADO'] == True].copy()
    else:
        df_repetidos = df_full[df_full.duplicated(subset=['CONTRATO'], keep=False)].copy()

    if df_repetidos.empty and 'CONTRATO' in df_full.columns:
        df_repetidos = df_full[df_full.duplicated(subset=['CONTRATO'], keep=False)].copy()

    return df_repetidos.drop_duplicates()


def gerar_contratos_por_banco(df_full: pd.DataFrame) -> pd.DataFrame:
    if df_full.empty:
        return pd.DataFrame()

    df = df_full.copy()
    if 'BANCO' not in df.columns:
        df['BANCO'] = 'NÃO INFORMADO'
    else:
        df['BANCO'] = df['BANCO'].fillna('NÃO INFORMADO')

    summary = (
        df.groupby('BANCO', dropna=False)
        .agg(
            TOTAL_CONTRATOS=('CONTRATO', 'size'),
            CONTRATOS_UNICOS=('CONTRATO', lambda s: s.nunique()),
            CONTRATOS_DUPLICADOS=('CONTRATO', lambda s: s.duplicated(keep=False).sum())
        )
        .reset_index()
    )
    return summary


def apply_excel_formatting(writer, df: pd.DataFrame, sheet_name: str):
    """
    Aplica formatação no Excel:
    - Colunas de data: formato DD/MM/YYYY
    - Coluna CONTRATO: formato texto
    """
    worksheet = writer.sheets[sheet_name]
    
    # Lista de possíveis nomes de colunas de data
    colunas_data_conhecidas = [
        'DT.ASS.', 'DT.EVENTO', 'DT.HAB.', 'DT.PROC.HAB.',
        'DT.ASS', 'DT.EVENTO', 'DT.HAB', 'DT.PROC.HAB',
        'DATA ASS.', 'DATA EVENTO', 'DATA HAB.', 'DATA PROC.HAB.',
        'DT.BASE', 'DT.TERM.ANALISE', 'DT.MANIFESTACAO', 'DT.POS.NOVACAO',
        'DT.ULT.AUDITORIA', 'DT.ULT.NEGOCIACAO', 'DATA STATUS'
    ]
    
    # Formatar colunas de data (nome, dtype datetime ou object com datas)
    for col_name in df.columns:
        is_dt_dtype = pd.api.types.is_datetime64_any_dtype(df[col_name])
        is_date_col = (
            col_name in colunas_data_conhecidas
            or str(col_name).upper().startswith('DT.')
            or str(col_name).upper().startswith('DATA')
            or is_dt_dtype
        )
        
        if is_date_col:
            try:
                col_idx = df.columns.get_loc(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"{col_letter}{row}"]
                    v = cell.value
                    if v is not None:
                        try:
                            if isinstance(v, datetime):
                                cell.value = v.date()
                            elif isinstance(v, date) and not isinstance(v, datetime):
                                cell.value = v
                            elif hasattr(v, 'to_pydatetime'):
                                pd_dt = v.to_pydatetime()
                                cell.value = pd_dt.date()
                        except Exception:
                            pass
                    cell.number_format = 'DD/MM/YYYY'
            except Exception as e:
                logger.warning(f"Erro ao formatar coluna de data {col_name}: {e}")
    
    # Formatar coluna CONTRATO como texto
    if 'CONTRATO' in df.columns:
        col_idx = df.columns.get_loc('CONTRATO') + 1
        col_letter = get_column_letter(col_idx)
        for row in range(2, len(df) + 2):
            cell = worksheet[f"{col_letter}{row}"]
            cell.number_format = '@'
    
    # Formatar coluna D (índice 3) como texto se existir
    if len(df.columns) > 3:
        col_idx = 4  # Coluna D (1-based)
        col_letter = get_column_letter(col_idx)
        for row in range(2, len(df) + 2):
            cell = worksheet[f"{col_letter}{row}"]
            cell.number_format = '@'


def add_column_ae_sum(writer, df: pd.DataFrame, sheet_name: str):
    """
    Adiciona soma da coluna AE na segunda linha abaixo do último valor.
    Coluna AE = índice 30 (0-based) ou 31 (1-based em Excel)
    """
    if len(df.columns) <= 30 or len(df) == 0:
        logger.debug(f"Não há coluna AE ou dados para adicionar soma em {sheet_name}")
        return
    
    try:
        worksheet = writer.sheets[sheet_name]
        
        # Coluna AE = índice 30 (0-based), que é coluna 31 no Excel (letra AE)
        col_ae = df.columns[30]
        col_letter = get_column_letter(31)  # AE
        
        # Linha onde colocar a soma (2 linhas abaixo do último dado)
        # Dados começam na linha 2, então última linha de dados = len(df) + 1
        # Soma vai 2 linhas abaixo = len(df) + 3
        row_sum = len(df) + 3
        
        # Calcular soma
        try:
            # Tentar converter coluna para numérico e somar
            valores_numericos = pd.to_numeric(df[col_ae], errors='coerce')
            soma = valores_numericos.sum()
            
            if pd.notna(soma):
                # Adicionar rótulo
                worksheet[f"AD{row_sum}"] = "SOMA AE:"
                # Adicionar valor da soma
                worksheet[f"{col_letter}{row_sum}"] = soma
                worksheet[f"{col_letter}{row_sum}"].number_format = '#,##0.00'
                
                logger.debug(f"Soma da coluna AE adicionada em {sheet_name}: {soma}")
        except Exception as e:
            logger.warning(f"Erro ao calcular soma da coluna AE: {e}")
            
    except Exception as e:
        logger.warning(f"Erro ao adicionar soma da coluna AE em {sheet_name}: {e}")


def save_processed_file(df: pd.DataFrame, filepath: str):
    """Salva arquivo Excel processado com formatação"""
    try:
        filepath_obj = Path(filepath)
        filepath_obj.parent.mkdir(parents=True, exist_ok=True)
        
        logger.debug(f"Salvando arquivo: {filepath}")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Dados', index=False)
            apply_excel_formatting(writer, df, 'Dados')
        
        if not filepath_obj.exists():
            raise Exception(f"Arquivo não foi salvo: {filepath}")
        
        logger.debug(f"Arquivo salvo com sucesso: {filepath}")
        
    except Exception as e:
        logger.error(f"Erro ao salvar arquivo {filepath}: {e}")
        raise


async def process_contratos(
    files: List[UploadFile], 
    bank_type: str, 
    filter_type: str = "todos", 
    file_type: str = "todos",
    period_filter_enabled: str = "false",
    reference_date: str = None,
    months_back: int = 2,
    habitacional_filter_enabled: str = "false",  # ✅ NOVO PARÂMETRO
    habitacional_reference_date: str = None,      # ✅ NOVO PARÂMETRO
    habitacional_months_back: int = 2             # ✅ NOVO PARÂMETRO
) -> StreamingResponse:
    """
    ✅ CORRIGIDO: Processa múltiplas planilhas Excel de contratos.
    Agora com suporte a filtro habitacional (colunas W e Y).
    
    Args:
        files: Lista de arquivos Excel
        bank_type: "bemge" ou "minas_caixa"
        filter_type: "auditado", "nauditado" ou "todos"
        file_type: "3026-11", "3026-12", "3026-15" ou "todos"
        period_filter_enabled: "true" ou "false" - Ativa filtro de período
        reference_date: Data de referência no formato "YYYY-MM-DD"
        months_back: Número de meses para trás (1, 2, 3, 4, 5, 6 ou 12)
        habitacional_filter_enabled: "true" ou "false" - Ativa filtro habitacional (NOVO)
        habitacional_reference_date: Data de referência para filtro habitacional (NOVO)
        habitacional_months_back: Número de meses para filtro habitacional (NOVO)
    
    Returns:
        StreamingResponse com arquivo Excel consolidado
    """
    try:
        logger.info(f"========================================")
        logger.info(f"INICIANDO PROCESSAMENTO DE CONTRATOS")
        logger.info(f"========================================")
        logger.info(f"Banco: {bank_type}")
        logger.info(f"Filtro: {filter_type}")
        logger.info(f"Tipo de arquivo: {file_type}")
        logger.info(f"Filtro de período: {period_filter_enabled}")
        logger.info(f"Data referência: {reference_date}")
        logger.info(f"Meses atrás: {months_back}")
        logger.info(f"✅ Filtro habitacional: {habitacional_filter_enabled}")  # NOVO
        logger.info(f"✅ Data referência habitacional: {habitacional_reference_date}")  # NOVO
        logger.info(f"✅ Meses atrás habitacional: {habitacional_months_back}")  # NOVO
        logger.info(f"Arquivos: {[f.filename for f in files]}")
        logger.info(f"========================================")
        
        # Validar bank_type
        bank_type_normalized = _normalize_bank_type_key(bank_type)
        if bank_type_normalized not in ['bemge', 'minas_caixa']:
            raise HTTPException(
                status_code=400,
                detail="bank_type deve ser 'bemge' ou 'minas_caixa'"
            )
        
        # Validar filter_type
        if filter_type not in ['auditado', 'nauditado', 'todos']:
            raise HTTPException(
                status_code=400,
                detail="filter_type deve ser 'auditado', 'nauditado' ou 'todos'"
            )
        
        # Validar file_type
        if file_type not in ['3026-11', '3026-12', '3026-15', 'todos']:
            raise HTTPException(
                status_code=400,
                detail="file_type deve ser '3026-11', '3026-12', '3026-15' ou 'todos'"
            )
        
        # Validar period_filter_enabled
        if period_filter_enabled not in ['true', 'false']:
            raise HTTPException(
                status_code=400,
                detail="period_filter_enabled deve ser 'true' ou 'false'"
            )
        
        # ✅ Validar habitacional_filter_enabled
        if habitacional_filter_enabled not in ['true', 'false']:
            raise HTTPException(
                status_code=400,
                detail="habitacional_filter_enabled deve ser 'true' ou 'false'"
            )
        
        # Converter para booleano
        period_filter_active = period_filter_enabled == "true"
        habitacional_filter_active = habitacional_filter_enabled == "true"  # ✅ NOVO
        
        # Validar se pelo menos um arquivo foi enviado
        if not files or len(files) == 0:
            raise HTTPException(
                status_code=400,
                detail="Pelo menos um arquivo deve ser enviado"
            )
        
        bank_name = get_bank_name(bank_type_normalized)
        sheet_names = get_sheet_names(bank_type_normalized)
        
        # Usar underscore em vez de espaços para nomes de pastas
        bank_folder = bank_type_normalized.replace(" ", "_")
        base_dir = Path(f"arquivo_morto/{bank_folder}")
        filtragem_dir = Path("arquivo_morto/3026 - Filtragens")
        
        # Criar estrutura de pastas
        try:
            base_dir.mkdir(parents=True, exist_ok=True)
            filtragem_dir.mkdir(parents=True, exist_ok=True)
            logger.debug(f"Pastas criadas: {base_dir}, {filtragem_dir}")
        except Exception as e:
            logger.error(f"Erro ao criar pastas: {e}")
            raise HTTPException(status_code=500, detail=f"Erro ao criar pastas: {str(e)}")
        
        # Estruturas para consolidar dados
        all_contratos = []
        dados_por_aba = {
            '3026-11': [],
            '3026-15': []
        }
        dados_3026_12 = {
            'todos': [],
            'auditados': [],
            'naud': [],
            'todos_ultimos_2_meses': [],
            'auditados_ultimos_2_meses': [],
            'naud_ultimos_2_meses': []
        }
        tem_3026_12 = False
        
        # Processar cada arquivo
        for file in files:
            filename = file.filename
            filename_upper = filename.upper()
            
            logger.info(f"\n{'='*60}")
            logger.info(f"📄 PROCESSANDO ARQUIVO: {filename}")
            logger.info(f"{'='*60}")
            
            # Filtrar por tipo de arquivo se especificado
            if file_type != "todos":
                file_type_normalized = file_type.upper().replace("-", "")
                if file_type.upper() not in filename_upper and file_type_normalized not in filename_upper:
                    logger.debug(f"Arquivo {filename} ignorado (não é {file_type})")
                    continue
            
            # Ler arquivo
            try:
                contents = await file.read()
                df = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
                logger.info(f"✅ Arquivo lido: {len(df)} linhas, {len(df.columns)} colunas")
            except Exception as e:
                logger.error(f"❌ Erro ao ler arquivo {filename}: {e}")
                raise HTTPException(
                    status_code=400,
                    detail=f"Erro ao ler arquivo {filename}: {str(e)}"
                )
            
            # Formatação de datas
            df = format_date_columns(df)
            df = format_object_columns_that_look_like_dates(df)
            
            # Remover colunas gerais
            df = remove_general_columns(df)
            
            # Formatação da coluna CONTRATO
            df = format_contrato_column(df)
            
            # Detectar tipo de arquivo
            detected_file_type = detect_file_type(filename)
            logger.info(f"Tipo detectado: {detected_file_type}")
            
            if detected_file_type == '3026-11':
                df_processado, total_linhas, total_unicos, total_duplicados = process_3026_11(df, bank_name)

                if df_processado.empty:
                    logger.warning(f"⚠️  Arquivo {filename} resultou em DataFrame vazio")
                    continue

                df_processado['TIPO_ARQUIVO'] = '3026-11'
                df_processado['BANCO'] = bank_name
                df_processado['DUPLICADO'] = df_processado['CONTRATO'].duplicated(keep=False)

                # Filtro por tipo (auditado/nauditado/todos)
                if filter_type != 'todos' and 'AUDITADO' in df_processado.columns:
                    df_processado['AUDITADO'] = df_processado['AUDITADO'].astype(str).str.upper().str.strip()
                    if filter_type == 'auditado':
                        df_processado = df_processado[
                            df_processado['AUDITADO'].isin(['AUDI', 'AUD', 'AUDITADO'])
                        ].copy()
                    elif filter_type == 'nauditado':
                        df_processado = df_processado[
                            df_processado['AUDITADO'].isin(['NAUD', 'NAO AUDITADO', 'NAUDITADO'])
                        ].copy()

                # ✅ APLICAR FILTRO HABITACIONAL (NOVO)
                if habitacional_filter_active:
                    logger.info(f"\n✅ APLICANDO FILTRO HABITACIONAL PARA 3026-11")
                    df_processado = filtrar_planilha_contratos(
                        df_processado,
                        aplicar_periodo=False,
                        aplicar_habitacional=True,
                        reference_date=habitacional_reference_date,
                        months_back=habitacional_months_back,
                        bank_type=bank_type_normalized
                    )
                    logger.info(f"✅ Filtro habitacional aplicado: {len(df_processado)} registros")

                # Filtro de período
                if period_filter_active:
                    logger.info(f"\n📅 APLICANDO FILTRO DE PERÍODO")
                    df_processado = filtrar_planilha_contratos(
                        df_processado,
                        aplicar_periodo=True,
                        reference_date=reference_date,
                        months_back=months_back,
                        bank_type=bank_type_normalized
                    )

                all_contratos.append(df_processado)
                dados_por_aba['3026-11'].append(df_processado.copy())

                n_arquivo = int(df_processado['CONTRATO'].nunique()) if 'CONTRATO' in df_processado.columns else len(df_processado)
                save_filename = f"3026-11 - {bank_name} - {n_arquivo} (CONTRATOS).xlsx"
                save_filepath = base_dir / save_filename
                save_processed_file(df_processado, str(save_filepath))

                if 'AUDITADO' in df_processado.columns:
                    filepath_filtragem = filtragem_dir / save_filename
                    save_processed_file(df_processado, str(filepath_filtragem))
                
                logger.info(f"✅ 3026-11 processado: {len(df_processado)} registros finais")

            elif detected_file_type == '3026-15':
                df_processado, total_linhas, total_unicos, total_duplicados = process_3026_15(df, bank_name)

                if df_processado.empty:
                    logger.warning(f"⚠️  Arquivo {filename} resultou em DataFrame vazio")
                    continue

                df_processado['TIPO_ARQUIVO'] = '3026-15'
                df_processado['BANCO'] = bank_name
                df_processado['DUPLICADO'] = df_processado['CONTRATO'].duplicated(keep=False)

                if filter_type != 'todos' and 'AUDITADO' in df_processado.columns:
                    df_processado['AUDITADO'] = df_processado['AUDITADO'].astype(str).str.upper().str.strip()
                    if filter_type == 'auditado':
                        df_processado = df_processado[
                            df_processado['AUDITADO'].isin(['AUDI', 'AUD', 'AUDITADO'])
                        ].copy()
                    elif filter_type == 'nauditado':
                        df_processado = df_processado[
                            df_processado['AUDITADO'].isin(['NAUD', 'NAO AUDITADO', 'NAUDITADO'])
                        ].copy()

                if period_filter_active:
                    logger.info(f"\n📅 APLICANDO FILTRO DE PERÍODO")
                    df_processado = filtrar_planilha_contratos(
                        df_processado,
                        aplicar_periodo=True,
                        reference_date=reference_date,
                        months_back=months_back,
                        bank_type=bank_type_normalized
                    )

                all_contratos.append(df_processado)
                dados_por_aba['3026-15'].append(df_processado.copy())

                n_arquivo = int(df_processado['CONTRATO'].nunique()) if 'CONTRATO' in df_processado.columns else len(df_processado)
                save_filename = f"3026-15 - {bank_name} - {n_arquivo} (CONTRATOS).xlsx"
                save_filepath = base_dir / save_filename
                save_processed_file(df_processado, str(save_filepath))

                if 'AUDITADO' in df_processado.columns:
                    filepath_filtragem = filtragem_dir / save_filename
                    save_processed_file(df_processado, str(filepath_filtragem))
                
                logger.info(f"✅ 3026-15 processado: {len(df_processado)} registros finais")

            elif detected_file_type == '3026-12':
                logger.info(f"🔧 Processando 3026-12...")
                
                resultados = processar_3026_12_com_abas(
                    df,
                    bank_name,
                    bank_type_normalized,
                    period_filter_active,
                    reference_date,
                    months_back
                )
                abas = resultados['abas']
                stats = resultados['stats']
                tem_3026_12 = True

                logger.info(f"📋 Abas disponíveis: {list(abas.keys())}")

                # Mapear nomes de chaves corretamente
                chaves_para_dados = {
                    'todos': 'todos',
                    'aud': 'auditados',  # ✅ Mapeamento correto
                    'naud': 'naud',
                    'auditados_ultimos_2_meses': 'auditados_ultimos_2_meses',
                    'naud_ultimos_2_meses': 'naud_ultimos_2_meses',
                    'todos_ultimos_2_meses': 'todos_ultimos_2_meses'
                }

                for chave_aba, chave_dados in chaves_para_dados.items():
                    if chave_aba in abas:
                        subset = abas[chave_aba]
                        if not subset.empty:
                            logger.info(f"   Adicionando {chave_dados}: {len(subset)} registros")
                            dados_3026_12[chave_dados].append(subset.copy())
                        else:
                            logger.debug(f"   {chave_dados} está vazio")

                # Processar AUD e NAUD para salvar arquivos individuais
                for tipo_label, subset_key_aba, subset_key_stats in [
                    ('AUD', 'aud', 'aud'), 
                    ('NAUD', 'naud', 'naud')
                ]:
                    df_subset = abas.get(subset_key_aba)
                    if df_subset is None or df_subset.empty:
                        logger.warning(f"   ⚠️  Subset {tipo_label} está vazio ou não existe")
                        continue

                    logger.info(f"   Processando {tipo_label}: {len(df_subset)} registros")

                    # Adicionar a all_contratos (sem filtro adicional para não duplicar)
                    all_contratos.append(df_subset.copy())

                    # Para salvar, aplicar filtro se necessário
                    df_para_salvar = df_subset.copy()
                    
                    if period_filter_active:
                        logger.info(f"      Aplicando filtro de período em {tipo_label}...")
                        df_para_salvar = filtrar_planilha_contratos(
                            df_para_salvar,
                            aplicar_periodo=True,
                            reference_date=reference_date,
                            months_back=months_back,
                            bank_type=bank_type_normalized
                        )
                        logger.info(f"      {tipo_label} após filtro: {len(df_para_salvar)} registros")

                    total_unicos = _stats_total_unicos(stats, subset_key_stats)
                    if total_unicos <= 0 and 'CONTRATO' in df_subset.columns:
                        total_unicos = int(df_subset['CONTRATO'].nunique())
                    save_filename = f"3026-12 - {bank_name} - {tipo_label} - {total_unicos} (CONTRATOS).xlsx"
                    save_filepath = base_dir / save_filename

                    if not df_para_salvar.empty:
                        logger.info(f"      Salvando {tipo_label}: {save_filename}")
                        save_processed_file(df_para_salvar, str(save_filepath))
                        filepath_filtragem = filtragem_dir / save_filename
                        save_processed_file(df_para_salvar, str(filepath_filtragem))
                    else:
                        logger.warning(f"      ⚠️  {tipo_label} vazio após filtragem, não salvando arquivo")
                
                logger.info(f"✅ 3026-12 processado com sucesso")
        
        # Consolidar todos os dados
        if not all_contratos:
            raise HTTPException(
                status_code=400,
                detail="Nenhum arquivo válido foi processado"
            )

        df_full = pd.concat(all_contratos, ignore_index=True)

        logger.info(f"\n{'='*60}")
        logger.info(f"📊 CONSOLIDAÇÃO FINAL")
        logger.info(f"{'='*60}")
        logger.info(f"Total de contratos consolidados: {len(df_full)}")

        df_filtrado = df_full.copy()
        if filter_type != 'todos':
            if 'AUDITADO_TIPO' in df_filtrado.columns:
                if filter_type == 'auditado':
                    df_filtrado = df_filtrado[df_filtrado['AUDITADO_TIPO'] == 'AUD'].copy()
                elif filter_type == 'nauditado':
                    df_filtrado = df_filtrado[df_filtrado['AUDITADO_TIPO'] == 'NAUD'].copy()
            elif 'AUDITADO' in df_filtrado.columns:
                df_filtrado['AUDITADO'] = df_filtrado['AUDITADO'].astype(str).str.upper().str.strip()
                if filter_type == 'auditado':
                    df_filtrado = df_filtrado[df_filtrado['AUDITADO'].isin(['AUDI', 'AUD', 'AUDITADO'])].copy()
                elif filter_type == 'nauditado':
                    df_filtrado = df_filtrado[df_filtrado['AUDITADO'].isin(['NAUD', 'NAO AUDITADO', 'NAUDITADO'])].copy()

        df_filtrado = filtrar_planilha_contratos(
            df_filtrado,
            aplicar_periodo=period_filter_active,
            reference_date=reference_date,
            months_back=months_back,
            bank_type=bank_type_normalized
        )

        # Mesma lógica da consolidação filtrada (inclui filter_type + período + DEST só no 3026-12)
        df_ultimos_2_meses = df_filtrado.copy() if period_filter_active else pd.DataFrame()

        df_resumo = gerar_resumo_geral(df_full)
        df_repetidos = gerar_contratos_repetidos(df_full)
        df_contratos_por_banco = gerar_contratos_por_banco(df_full)
        
        logger.info(f"✅ Dados filtrados: {len(df_filtrado)} registros")
        logger.info(f"✅ Resumos gerados com sucesso")
        
        # Criar arquivo Excel consolidado
        output = io.BytesIO()

        logger.info(f"\n📝 Criando arquivo Excel consolidado...")

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            if not df_resumo.empty:
                df_resumo.to_excel(writer, sheet_name='Resumo Geral', index=False)
                apply_excel_formatting(writer, df_resumo, 'Resumo Geral')
            else:
                pd.DataFrame({'Mensagem': ['Resumo geral não disponível']}).to_excel(
                    writer, sheet_name='Resumo Geral', index=False
                )

            nome_repetidos = 'Contratos Repetidos'
            if not df_repetidos.empty:
                df_repetidos.to_excel(writer, sheet_name=nome_repetidos, index=False)
                apply_excel_formatting(writer, df_repetidos, nome_repetidos)
            else:
                pd.DataFrame({'Mensagem': ['Nenhum contrato repetido encontrado']}).to_excel(
                    writer, sheet_name=nome_repetidos, index=False
                )

            nome_por_banco = 'Contratos por Banco'
            if not df_contratos_por_banco.empty:
                df_contratos_por_banco.to_excel(writer, sheet_name=nome_por_banco, index=False)
                apply_excel_formatting(writer, df_contratos_por_banco, nome_por_banco)
            else:
                pd.DataFrame({'Mensagem': ['Nenhum contrato por banco encontrado']}).to_excel(
                    writer, sheet_name=nome_por_banco, index=False
                )

            if dados_por_aba['3026-11']:
                df_3026_11 = pd.concat(dados_por_aba['3026-11'], ignore_index=True)
                nome_aba_11 = sheet_names['3026-11'][:31]
                df_3026_11.to_excel(writer, sheet_name=nome_aba_11, index=False)
                apply_excel_formatting(writer, df_3026_11, nome_aba_11)

            if tem_3026_12:
                nome_12_todos = sheet_names['3026-12-TODOS'][:31]
                if dados_3026_12['todos']:
                    df_3026_12_todos = pd.concat(dados_3026_12['todos'], ignore_index=True)
                    df_3026_12_todos.to_excel(writer, sheet_name=nome_12_todos, index=False)
                    apply_excel_formatting(writer, df_3026_12_todos, nome_12_todos)

                if dados_3026_12['auditados']:
                    df_3026_12_aud = pd.concat(dados_3026_12['auditados'], ignore_index=True)
                    nome_aba_12_aud = sheet_names['3026-12-AUD'][:31]
                    df_3026_12_aud.to_excel(writer, sheet_name=nome_aba_12_aud, index=False)
                    apply_excel_formatting(writer, df_3026_12_aud, nome_aba_12_aud)
                    add_column_ae_sum(writer, df_3026_12_aud, nome_aba_12_aud)

                if dados_3026_12['naud']:
                    df_3026_12_naud = pd.concat(dados_3026_12['naud'], ignore_index=True)
                    nome_aba_12_naud = sheet_names['3026-12-NAUD'][:31]
                    df_3026_12_naud.to_excel(writer, sheet_name=nome_aba_12_naud, index=False)
                    apply_excel_formatting(writer, df_3026_12_naud, nome_aba_12_naud)
                    add_column_ae_sum(writer, df_3026_12_naud, nome_aba_12_naud)

                if period_filter_active:
                    for chave, nome_chave in [
                        ('auditados_ultimos_2_meses', sheet_names['3026-12-ULTIMOS_AUD']),
                        ('naud_ultimos_2_meses', sheet_names['3026-12-ULTIMOS_NAUD']),
                        ('todos_ultimos_2_meses', sheet_names['3026-12-ULTIMOS_TODOS'])
                    ]:
                        if dados_3026_12[chave]:
                            df_periodo = pd.concat(dados_3026_12[chave], ignore_index=True)
                            nome_periodo = nome_chave[:31]
                            df_periodo.to_excel(writer, sheet_name=nome_periodo, index=False)
                            apply_excel_formatting(writer, df_periodo, nome_periodo)

            if not df_filtrado.empty:
                df_filtrado.to_excel(writer, sheet_name='Dados Filtrados', index=False)
                apply_excel_formatting(writer, df_filtrado, 'Dados Filtrados')
            else:
                pd.DataFrame({'Mensagem': ['Filtros removeram todos os contratos']}).to_excel(
                    writer, sheet_name='Dados Filtrados', index=False
                )

            if dados_por_aba['3026-15']:
                df_3026_15 = pd.concat(dados_por_aba['3026-15'], ignore_index=True)
                nome_aba_15 = sheet_names['3026-15'][:31]
                df_3026_15.to_excel(writer, sheet_name=nome_aba_15, index=False)
                apply_excel_formatting(writer, df_3026_15, nome_aba_15)

            nome_aba_periodo = f'Últimos {months_back} Meses'[:31]
            if period_filter_active:
                if not df_ultimos_2_meses.empty:
                    df_ultimos_2_meses.to_excel(writer, sheet_name=nome_aba_periodo, index=False)
                    apply_excel_formatting(writer, df_ultimos_2_meses, nome_aba_periodo)
                else:
                    pd.DataFrame({'Mensagem': [f'Nenhum contrato encontrado nos últimos {months_back} meses']}).to_excel(
                        writer, sheet_name=nome_aba_periodo, index=False
                    )
            else:
                pd.DataFrame({'Mensagem': ['Filtro de período desativado']}).to_excel(
                    writer, sheet_name=nome_aba_periodo, index=False
                )
        
        # Resetar ponteiro e ler dados
        output.seek(0)
        excel_data = output.read()
        output.close()
        
        # Nome do arquivo de saída - varia com base no file_type
        filtro_nome = filter_type.upper()
        banco_nome = "BEMGE" if bank_type_normalized == "bemge" else "MINAS_CAIXA"
        
        # Variar nome do arquivo baseado no file_type
        if file_type == "todos":
            tipo_nome = "TODOS_TIPOS"
        elif file_type == "3026-11":
            tipo_nome = "3026-11"
        elif file_type == "3026-12":
            tipo_nome = "3026-12"
        elif file_type == "3026-15":
            tipo_nome = "3026-15"
        else:
            tipo_nome = file_type.upper().replace("-", "_")
        
        periodo_nome = f"_{months_back}MESES" if period_filter_active else ""
        habitacional_nome = "_HABITACIONAL" if habitacional_filter_active else ""  # ✅ NOVO
        
        filename_output = f"contratos_{tipo_nome}_{banco_nome}_{filtro_nome}{periodo_nome}{habitacional_nome}_consolidado.xlsx"
        
        logger.info(f"\n{'='*60}")
        logger.info(f"✅ PROCESSAMENTO CONCLUÍDO COM SUCESSO")
        logger.info(f"{'='*60}")
        logger.info(f"Arquivo gerado: {filename_output}")
        logger.info(f"Total de registros finais: {len(df_filtrado)}")
        logger.info(f"{'='*60}\n")
        
        return StreamingResponse(
            io.BytesIO(excel_data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename_output}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ ERRO CRÍTICO ao processar contratos: {e}")
        logger.error(f"Tipo do erro: {type(e).__name__}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao processar contratos: {str(e)}"
        )
